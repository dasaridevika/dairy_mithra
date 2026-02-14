from flask import Flask, render_template, request, redirect, session, send_from_directory
from openpyxl import Workbook, load_workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import TableStyle
import os

app = Flask(__name__)
app.secret_key = "dairymithra_secret"

PRICE_PER_LITRE = 50
DAYS_IN_MONTH = 30

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXCEL_FILE = os.path.join(UPLOAD_FOLDER, "customers.xlsx")


# ------------------ EXCEL SETUP ------------------

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["phone","name","role","milk","delivery","totalAmount","payments"])
        wb.save(EXCEL_FILE)


def get_all_users():
    initialize_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        users.append({
            "phone": row[0],
            "name": row[1],
            "role": row[2],
            "milk": row[3],
            "delivery": row[4],
            "totalAmount": row[5],
            "payments": row[6]
        })
    return users


def save_user(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data["phone"],
        data["name"],
        data["role"],
        data["milk"],
        data["delivery"],
        data["totalAmount"],
        data["payments"]
    ])
    wb.save(EXCEL_FILE)


# ------------------ ROUTES ------------------

@app.route("/")
def home():
    if "user" in session:
        if session["role"] == "owner":
            return redirect("/owner")
        else:
            return redirect("/customer")
    return render_template("index.html")


@app.route("/login", methods=["POST"])
def login():
    phone = request.form.get("phone")
    users = get_all_users()
    user = next((u for u in users if u["phone"] == phone), None)

    if user:
        session["user"] = phone
        session["role"] = user["role"]
        return redirect("/owner" if user["role"]=="owner" else "/customer")

    session["pending_phone"] = phone
    return redirect("/role")


@app.route("/role", methods=["GET","POST"])
def role():
    if request.method == "POST":
        role = request.form.get("role")
        phone = session.get("pending_phone")

        data = {
            "phone": phone,
            "name": "User_" + phone[-4:],
            "role": role,
            "milk": 1,
            "delivery": True,
            "totalAmount": 0,
            "payments": 0
        }

        save_user(data)

        session["user"] = phone
        session["role"] = role

        return redirect("/owner" if role=="owner" else "/customer")

    return render_template("role.html")


@app.route("/owner")
def owner_dashboard():
    if "user" not in session or session["role"] != "owner":
        return redirect("/")

    users = get_all_users()

    for u in users:
        if u["role"] == "customer":
            milk = float(u["milk"] or 0)
            u["totalAmount"] = milk * PRICE_PER_LITRE * DAYS_IN_MONTH

    return render_template("ownerdashboard.html", users=users)


@app.route("/customer")
def customer_dashboard():
    if "user" not in session or session["role"] != "customer":
        return redirect("/")

    users = get_all_users()
    user = next((u for u in users if u["phone"] == session["user"]), None)

    milk = float(user["milk"] or 0)
    user["totalAmount"] = milk * PRICE_PER_LITRE * DAYS_IN_MONTH

    return render_template("customer.html", user=user)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# ------------------ ADD CUSTOMER ------------------

@app.route("/add_customer", methods=["POST"])
def add_customer():
    name = request.form["name"]
    phone = request.form["phone"]

    data = {
        "phone": phone,
        "name": name,
        "role": "customer",
        "milk": 1,
        "delivery": True,
        "totalAmount": 0,
        "payments": 0
    }

    save_user(data)
    return redirect("/owner")


# ------------------ UPDATE MILK ------------------

@app.route("/update_user", methods=["POST"])
def update_user():
    phone = request.form.get("phone")
    milk = float(request.form.get("milk"))

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == phone:
            row[3].value = milk
            row[5].value = milk * PRICE_PER_LITRE * DAYS_IN_MONTH
            break

    wb.save(EXCEL_FILE)
    return redirect("/owner")


# ------------------ ADD PAYMENT ------------------

@app.route("/add_payment", methods=["POST"])
def add_payment():
    phone = request.form.get("phone")
    payment = float(request.form.get("payment"))

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == phone:
            row[6].value = float(row[6].value or 0) + payment
            break

    wb.save(EXCEL_FILE)
    return redirect("/owner")


# ------------------ DELETE ------------------

@app.route("/delete_user/<phone>")
def delete_user(phone):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == phone:
            ws.delete_rows(row[0].row, 1)
            break

    wb.save(EXCEL_FILE)
    return redirect("/owner")


# ------------------ DELIVERY TOGGLE ------------------

@app.route("/toggle_delivery/<phone>")
def toggle_delivery(phone):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == phone:
            row[4].value = not row[4].value
            break

    wb.save(EXCEL_FILE)
    return redirect("/owner")


# ------------------ PDF INVOICE ------------------

@app.route("/download_pdf/<phone>")
def download_pdf(phone):
    users = get_all_users()
    user = next((u for u in users if u["phone"] == phone), None)

    file_path = os.path.join(UPLOAD_FOLDER, f"{phone}_invoice.pdf")
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    milk = float(user["milk"] or 0)
    total = milk * PRICE_PER_LITRE * DAYS_IN_MONTH
    payments = float(user["payments"] or 0)
    pending = total - payments

    data = [
        ["Customer", user["name"]],
        ["Phone", user["phone"]],
        ["Milk/Day", milk],
        ["Total", total],
        ["Payments", payments],
        ["Pending", pending]
    ]

    table = Table(data)
    table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.black)]))

    elements.append(Paragraph("DairyMithra Invoice", styles["Title"]))
    elements.append(Spacer(1,20))
    elements.append(table)
    doc.build(elements)

    return send_from_directory(UPLOAD_FOLDER, f"{phone}_invoice.pdf")


if __name__ == "__main__":
    app.run(debug=True)
